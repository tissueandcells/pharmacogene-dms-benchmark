"""
Generate Supplementary Tables S1, S2, S3 for the PharmEpi benchmark paper.

Outputs to results/supplementary/:
  - Supplementary_Table_S1_saturation_profile.xlsx
  - Supplementary_Table_S2_CYP2C9_function_specific.xlsx
  - Supplementary_Table_S3_NUDT15_substrate_specific.xlsx

Each file includes a README sheet describing its contents, columns, and
category thresholds used in the paper.

Category thresholds (identical to paper Section 3.6):
  CYP2C9 (activity, abundance):
    concordant WT-like    : activity > 0.7 AND abundance > 0.7
    concordant LOF        : activity < 0.3 AND abundance < 0.3
    stable-but-dead       : abundance > 0.7 AND activity  < 0.3
    unstable-but-active   : abundance < 0.3 AND activity  > 0.7
    intermediate          : otherwise
  NUDT15 (abundance, sensitivity):
    concordant-resistant  : abundance > 0.7 AND sensitivity > 0.7
    concordant-sensitive  : abundance < 0.3 AND sensitivity < 0.5
    thio-sensitive-stable : abundance > 0.7 AND sensitivity < 0.5
    paradoxical-resistant : abundance < 0.3 AND sensitivity > 0.8
    intermediate          : otherwise

Usage:
    python scripts/27_generate_supplementary.py
"""

import re
import gzip
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
OUT = Path("results/supplementary")
OUT.mkdir(parents=True, exist_ok=True)

DMS_PATHS = {
    "CYP2C9 activity":    "data/raw/cyp2c9/cyp2c9_activity_amorosi2021.csv",
    "CYP2C9 abundance":   "data/raw/cyp2c9/cyp2c9_abundance_amorosi2021.csv",
    "CYP2C19 abundance":  "data/raw/cyp2c19/cyp2c19_abundance_boyle2024.csv",
    "NUDT15 stability":   "data/raw/nudt15/nudt15_stability_suiter2020.csv",
    "NUDT15 activity":    "data/raw/nudt15/nudt15_activity_suiter2020.csv",
}
NUDT15_COMBINED = "data/raw/nudt15/nudt15_combined_suiter2020.csv"
AM_PATH = Path("data/external/AlphaMissense_aa_substitutions.tsv.gz")


# -----------------------------------------------------------------------------
# HGVS parsing
# -----------------------------------------------------------------------------
AA_3TO1 = {
    "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C",
    "Gln": "Q", "Glu": "E", "Gly": "G", "His": "H", "Ile": "I",
    "Leu": "L", "Lys": "K", "Met": "M", "Phe": "F", "Pro": "P",
    "Ser": "S", "Thr": "T", "Trp": "W", "Tyr": "Y", "Val": "V",
}
HGVS_RE = re.compile(r"^p\.([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})$")


def parse_hgvs(s):
    """Parse 'p.Ala123Val' -> ('A', 123, 'V'). Returns None for non-missense."""
    if not isinstance(s, str):
        return None
    m = HGVS_RE.match(s.strip())
    if not m:
        return None
    wt3, pos, mut3 = m.group(1), int(m.group(2)), m.group(3)
    wt = AA_3TO1.get(wt3)
    mut = AA_3TO1.get(mut3)
    if wt is None or mut is None:
        return None
    return (wt, pos, mut)


def autodetect_hgvs_col(df):
    for c in df.columns:
        if c.lower() == "hgvs_pro":
            return c
    for c in df.columns:
        if "hgvs" in c.lower() and "pro" in c.lower():
            return c
    return None


def autodetect_score_col(df, hint=""):
    hint = hint.lower()
    if "activity" in hint:
        prefs = ["activity_score", "score"]
    elif "abundance" in hint or "stability" in hint:
        prefs = ["abundance_score", "stability_score", "score"]
    elif "sensitivity" in hint:
        prefs = ["sensitivity_score", "score"]
    else:
        prefs = ["score"]
    for p in prefs:
        for c in df.columns:
            if c.lower() == p:
                return c
    return None


def load_dms(path, hint=""):
    """Load a DMS CSV, parse HGVS, return standardized dataframe."""
    df = pd.read_csv(path)
    hgvs_col = autodetect_hgvs_col(df)
    score_col = autodetect_score_col(df, hint)
    if hgvs_col is None or score_col is None:
        raise ValueError(
            f"Missing columns in {path}: hgvs={hgvs_col}, score={score_col}"
        )
    print(f"  [load] {Path(path).name}: hgvs={hgvs_col}, score={score_col}, n={len(df):,}")
    parsed = df[hgvs_col].apply(parse_hgvs)
    df = df.loc[parsed.notna()].copy()
    parsed = parsed[parsed.notna()]
    df["WT"] = [p[0] for p in parsed]
    df["pos"] = [p[1] for p in parsed]
    df["mut"] = [p[2] for p in parsed]
    df["hgvs_pro"] = df[hgvs_col].astype(str)
    df["_score"] = pd.to_numeric(df[score_col], errors="coerce")
    out = df[["hgvs_pro", "WT", "pos", "mut", "_score"]].dropna(subset=["_score"])
    return out


# -----------------------------------------------------------------------------
# AlphaMissense
# -----------------------------------------------------------------------------
def load_am():
    """Load AlphaMissense scores for the three pharmacogenes of interest."""
    if not AM_PATH.exists():
        print(f"[AM] File not found: {AM_PATH}")
        return {}
    print(f"[AM] Loading {AM_PATH}...")
    wanted = {"P11712", "P33261", "Q9NV35"}
    rows = {u: [] for u in wanted}
    with gzip.open(AM_PATH, "rt") as f:
        header = None
        for line in f:
            if line.startswith("#"):
                continue
            if header is None:
                header = line.strip().split("\t")
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 4:
                continue
            u = parts[0]
            if u not in wanted:
                continue
            var = parts[1]
            if len(var) < 3:
                continue
            rows[u].append((var[0], int(var[1:-1]), var[-1], float(parts[2]), parts[3]))
    out = {}
    for u, r in rows.items():
        out[u] = pd.DataFrame(r, columns=["WT", "pos", "mut", "am_score", "am_class"])
        print(f"  [AM] {u}: {len(out[u]):,} variants")
    return out


# -----------------------------------------------------------------------------
# Category assignment (identical to paper Section 3.6)
# -----------------------------------------------------------------------------
def categorize_cyp2c9(row):
    a = row.activity_score
    b = row.abundance_score
    if a > 0.7 and b > 0.7:
        return "concordant WT-like"
    if a < 0.3 and b < 0.3:
        return "concordant LOF"
    if a < 0.3 and b > 0.7:
        return "stable-but-dead"
    if a > 0.7 and b < 0.3:
        return "unstable-but-active"
    return "intermediate"


def categorize_nudt15(row):
    a = row.abundance_score
    s = row.sensitivity_score
    if a > 0.7 and s > 0.7:
        return "concordant-resistant"
    if a < 0.3 and s < 0.5:
        return "concordant-sensitive"
    if a > 0.7 and s < 0.5:
        return "thio-sensitive-stable"
    if a < 0.3 and s > 0.8:
        return "paradoxical-resistant"
    return "intermediate"


# -----------------------------------------------------------------------------
# Formatting helpers
# -----------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
HEADER_FONT = Font(name="Arial", size=11, bold=True)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=12, bold=True)


def format_sheet(ws, freeze_header=True):
    """Apply consistent formatting: fonts, header fill, column widths, frozen panes."""
    # Header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Body rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    # Column widths based on header length + small margin
    for col_idx, cell in enumerate(ws[1], start=1):
        header_len = len(str(cell.value)) if cell.value else 10
        # Check a few sample rows for max content width
        max_len = header_len
        for r in range(2, min(ws.max_row + 1, 20)):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)
    if freeze_header:
        ws.freeze_panes = "A2"


def write_readme(ws, title, lines):
    """Write a human-readable README sheet."""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 110
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line).font = BODY_FONT
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def finalize_workbook(path, readme_title, readme_lines):
    """Re-open the workbook, insert a README sheet at position 0, apply formatting."""
    wb = load_workbook(path)
    # Remove any pre-existing README
    if "README" in wb.sheetnames:
        del wb["README"]
    readme = wb.create_sheet("README", 0)
    write_readme(readme, readme_title, readme_lines)
    # Format data sheets
    for name in wb.sheetnames:
        if name == "README":
            continue
        format_sheet(wb[name])
    wb.save(path)


# -----------------------------------------------------------------------------
# Supplementary Table S1: per-position saturation profile
# -----------------------------------------------------------------------------
def build_s1():
    print("\n[S1] Building per-position saturation profile...")
    all_rows = []
    for label, path in DMS_PATHS.items():
        p = Path(path)
        if not p.exists():
            print(f"  [WARN] missing: {path} — skipping {label}")
            continue
        df = load_dms(p, hint=label)
        cov = (
            df.groupby("pos")
              .agg(WT=("WT", "first"), n_measured=("mut", "nunique"))
              .reset_index()
        )
        cov["dataset"] = label
        cov["n_possible"] = 19
        cov["saturation_pct"] = (cov["n_measured"] / 19 * 100).round(1)
        cov["fully_saturated"] = cov["n_measured"] == 19
        all_rows.append(
            cov[["dataset", "pos", "WT", "n_measured",
                 "n_possible", "saturation_pct", "fully_saturated"]]
        )

    if not all_rows:
        raise RuntimeError("No DMS files loaded. Check DMS_PATHS.")

    combined = pd.concat(all_rows, ignore_index=True)
    summary = (
        combined.groupby("dataset")
                .agg(n_positions=("pos", "nunique"),
                     mean_saturation_pct=("saturation_pct", "mean"),
                     median_saturation_pct=("saturation_pct", "median"),
                     n_fully_saturated=("fully_saturated", "sum"))
                .round(2)
                .reset_index()
    )
    # Reorder summary to match paper Table 1 row order
    ordered = ["CYP2C9 activity", "CYP2C9 abundance", "CYP2C19 abundance",
               "NUDT15 stability", "NUDT15 activity"]
    summary["dataset"] = pd.Categorical(summary["dataset"], categories=ordered, ordered=True)
    summary = summary.sort_values("dataset").reset_index(drop=True)

    out_path = OUT / "Supplementary_Table_S1_saturation_profile.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        combined.to_excel(w, sheet_name="Per-position saturation", index=False)
        summary.to_excel(w, sheet_name="Summary", index=False)

    readme_lines = [
        "Content: Per-residue saturation profile of the five deep mutational scanning (DMS) datasets analyzed in this study.",
        "",
        "Motivation: Quantifies the experimental coverage of each DMS assay across the protein sequence, enabling readers to identify positions and regions where coverage is sparse and interpret downstream model performance accordingly.",
        "",
        "Referenced in paper: Section 3.1 (Dataset characterization and coverage), Table 1 summary counts.",
        "",
        "Sheets:",
        "  1) 'Per-position saturation' — one row per (dataset, residue position).",
        "       dataset          : Name of the DMS dataset.",
        "       pos              : Residue position (1-based) in the canonical UniProt sequence.",
        "       WT               : Wild-type single-letter amino acid at this position.",
        "       n_measured       : Number of unique non-WT amino-acid substitutions with a reported score.",
        "       n_possible       : Maximum number of possible non-WT substitutions (always 19).",
        "       saturation_pct   : 100 * n_measured / 19.",
        "       fully_saturated  : True iff n_measured == 19.",
        "",
        "  2) 'Summary' — one row per DMS dataset.",
        "       n_positions           : Number of distinct residue positions with at least one measurement.",
        "       mean_saturation_pct   : Mean saturation percentage across positions.",
        "       median_saturation_pct : Median saturation percentage across positions.",
        "       n_fully_saturated     : Number of positions with all 19 non-WT substitutions measured.",
        "",
        "Data sources:",
        "  CYP2C9 activity, abundance : Amorosi et al. (2021), MaveDB urn:mavedb:00000095-a-1 and 00000095-b-1.",
        "  CYP2C19 abundance          : Boyle et al. (2024), MaveDB urn:mavedb:00001199-a-1.",
        "  NUDT15 stability, activity : Suiter et al. (2020), MaveDB urn:mavedb:00000055-a-1 and 00000055-b-1.",
    ]
    finalize_workbook(out_path, "Supplementary Table S1 — Per-position saturation profile", readme_lines)
    print(f"[S1] Saved: {out_path}  ({len(combined):,} position rows)")


# -----------------------------------------------------------------------------
# Supplementary Table S2: CYP2C9 paired variants + function-specific subset
# -----------------------------------------------------------------------------
def build_s2(am_lookup):
    print("\n[S2] Building CYP2C9 paired variant table...")
    act = load_dms(DMS_PATHS["CYP2C9 activity"], hint="activity").rename(
        columns={"_score": "activity_score"}
    )
    abu = load_dms(DMS_PATHS["CYP2C9 abundance"], hint="abundance").rename(
        columns={"_score": "abundance_score"}
    )
    all_paired = act.merge(
        abu[["WT", "pos", "mut", "abundance_score"]],
        on=["WT", "pos", "mut"], how="inner",
    )
    print(f"  All paired variants: {len(all_paired):,}")

    all_paired["delta"] = all_paired["activity_score"] - all_paired["abundance_score"]
    all_paired["abs_delta"] = all_paired["delta"].abs()
    all_paired["category"] = all_paired.apply(categorize_cyp2c9, axis=1)

    if "P11712" in am_lookup:
        all_paired = all_paired.merge(
            am_lookup["P11712"], on=["WT", "pos", "mut"], how="left"
        )
    else:
        all_paired["am_score"] = np.nan
        all_paired["am_class"] = ""

    all_paired = all_paired[[
        "hgvs_pro", "WT", "pos", "mut",
        "activity_score", "abundance_score", "delta", "abs_delta",
        "category", "am_score", "am_class",
    ]]
    all_paired = all_paired.sort_values(
        ["category", "abs_delta"], ascending=[True, False]
    ).reset_index(drop=True)

    fs = (all_paired[all_paired["abs_delta"] > 0.3]
          .sort_values("abs_delta", ascending=False)
          .reset_index(drop=True))
    print(f"  Function-specific subset (|delta| > 0.3): {len(fs):,}")

    cat_sum = (all_paired.groupby("category")
                         .agg(N=("hgvs_pro", "count"),
                              mean_activity=("activity_score", "mean"),
                              mean_abundance=("abundance_score", "mean"),
                              mean_am_score=("am_score", "mean"))
                         .round(3).reset_index())
    am_brk = (all_paired.groupby(["category", "am_class"])
                        .size().unstack(fill_value=0).reset_index())

    out_path = OUT / "Supplementary_Table_S2_CYP2C9_function_specific.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        all_paired.to_excel(w, sheet_name="All paired variants", index=False)
        fs.to_excel(w, sheet_name="Function-specific subset", index=False)
        cat_sum.to_excel(w, sheet_name="Category summary", index=False)
        am_brk.to_excel(w, sheet_name="AM class per category", index=False)

    readme_lines = [
        "Content: All CYP2C9 variants with paired activity (Click-seq) and abundance (VAMP-seq) measurements from Amorosi et al. (2021), together with AlphaMissense pathogenicity annotations.",
        "",
        f"Total paired variants: {len(all_paired):,} (activity set intersected with abundance set on HGVS identity).",
        f"Function-specific subset: {len(fs):,} variants with |activity - abundance| > 0.3 normalized units.",
        "",
        "Referenced in paper: Section 3.2 (functional heterogeneity), Section 3.6 (AlphaMissense stratified analysis), Figure 1 Panel B, Figure 2 Panels A–B, Table 4.",
        "",
        "Category thresholds (matching paper Section 3.6):",
        "  concordant WT-like   : activity > 0.7 AND abundance > 0.7",
        "  concordant LOF       : activity < 0.3 AND abundance < 0.3",
        "  stable-but-dead      : abundance > 0.7 AND activity  < 0.3",
        "  unstable-but-active  : abundance < 0.3 AND activity  > 0.7",
        "  intermediate         : otherwise",
        "",
        "Sheets:",
        "  1) 'All paired variants' — complete joint set, 11 columns:",
        "       hgvs_pro          : HGVS protein-level variant identifier (e.g., p.Ala123Val).",
        "       WT, pos, mut      : Wild-type AA, 1-based position, mutant AA.",
        "       activity_score    : Normalized Click-seq catalytic activity (Amorosi et al., 2021).",
        "       abundance_score   : Normalized VAMP-seq protein abundance (Amorosi et al., 2021).",
        "       delta             : activity_score minus abundance_score.",
        "       abs_delta         : |delta|, used as the function-specificity magnitude.",
        "       category          : Biochemical category (see thresholds above).",
        "       am_score          : AlphaMissense pathogenicity score, 0 (benign) to 1 (pathogenic).",
        "       am_class          : AlphaMissense discrete classification (benign / ambiguous / pathogenic).",
        "",
        "  2) 'Function-specific subset' — same columns, restricted to rows with abs_delta > 0.3.",
        "  3) 'Category summary' — per-category N, mean activity, mean abundance, mean AM score.",
        "  4) 'AM class per category' — cross-tabulation of category x AlphaMissense class counts.",
        "",
        "Data sources:",
        "  DMS: Amorosi C.J., et al. (2021), MaveDB urn:mavedb:00000095-a-1 (activity) and 00000095-b-1 (abundance).",
        "  AlphaMissense: Cheng J., et al. (2023), precomputed scores for human canonical sequences (UniProt P11712).",
    ]
    finalize_workbook(out_path, "Supplementary Table S2 — CYP2C9 paired variants with function-specific annotation", readme_lines)
    print(f"[S2] Saved: {out_path}  (all={len(all_paired):,}, function-specific={len(fs):,})")


# -----------------------------------------------------------------------------
# Supplementary Table S3: NUDT15 paired variants + substrate-specific subset
# -----------------------------------------------------------------------------
def build_s3(am_lookup):
    print("\n[S3] Building NUDT15 paired variant table...")
    comb_p = Path(NUDT15_COMBINED)
    df = None
    if comb_p.exists():
        raw = pd.read_csv(comb_p)
        hgvs_col = autodetect_hgvs_col(raw)
        if (hgvs_col
                and "abundance_score" in raw.columns
                and "sensitivity_score" in raw.columns):
            parsed = raw[hgvs_col].apply(parse_hgvs)
            raw = raw.loc[parsed.notna()].copy()
            parsed = parsed[parsed.notna()]
            raw["WT"] = [p[0] for p in parsed]
            raw["pos"] = [p[1] for p in parsed]
            raw["mut"] = [p[2] for p in parsed]
            df = raw.rename(columns={hgvs_col: "hgvs_pro"})[[
                "hgvs_pro", "WT", "pos", "mut",
                "abundance_score", "sensitivity_score",
            ]]
            df = df.dropna(subset=["abundance_score", "sensitivity_score"])
            print(f"  Using combined NUDT15 file: {len(df):,} paired variants")

    if df is None:
        print("  Combined file not available; merging stability + activity datasets...")
        stab = load_dms(DMS_PATHS["NUDT15 stability"], hint="stability").rename(
            columns={"_score": "abundance_score"}
        )
        act = load_dms(DMS_PATHS["NUDT15 activity"], hint="sensitivity").rename(
            columns={"_score": "sensitivity_score"}
        )
        df = stab.merge(
            act[["WT", "pos", "mut", "sensitivity_score"]],
            on=["WT", "pos", "mut"], how="inner",
        )
        df = df[[
            "hgvs_pro", "WT", "pos", "mut",
            "abundance_score", "sensitivity_score",
        ]]

    df["delta"] = df["sensitivity_score"] - df["abundance_score"]
    df["abs_delta"] = df["delta"].abs()
    df["category"] = df.apply(categorize_nudt15, axis=1)

    if "Q9NV35" in am_lookup:
        df = df.merge(am_lookup["Q9NV35"], on=["WT", "pos", "mut"], how="left")
    else:
        df["am_score"] = np.nan
        df["am_class"] = ""

    df = df[[
        "hgvs_pro", "WT", "pos", "mut",
        "abundance_score", "sensitivity_score", "delta", "abs_delta",
        "category", "am_score", "am_class",
    ]]
    df = df.sort_values(
        ["category", "abs_delta"], ascending=[True, False]
    ).reset_index(drop=True)

    ss = (df[df["abs_delta"] > 0.3]
          .sort_values("abs_delta", ascending=False)
          .reset_index(drop=True))
    print(f"  Substrate-specific subset (|delta| > 0.3): {len(ss):,}")

    cat_sum = (df.groupby("category")
                 .agg(N=("hgvs_pro", "count"),
                      mean_abundance=("abundance_score", "mean"),
                      mean_sensitivity=("sensitivity_score", "mean"),
                      mean_am_score=("am_score", "mean"))
                 .round(3).reset_index())
    am_brk = (df.groupby(["category", "am_class"])
                .size().unstack(fill_value=0).reset_index())

    out_path = OUT / "Supplementary_Table_S3_NUDT15_substrate_specific.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="All paired variants", index=False)
        ss.to_excel(w, sheet_name="Substrate-specific subset", index=False)
        cat_sum.to_excel(w, sheet_name="Category summary", index=False)
        am_brk.to_excel(w, sheet_name="AM class per category", index=False)

    readme_lines = [
        "Content: All NUDT15 variants with paired abundance (VAMP-seq) and thiopurine sensitivity measurements from Suiter et al. (2020), together with AlphaMissense pathogenicity annotations.",
        "",
        f"Total paired variants: {len(df):,} (from the combined dataset with both abundance_score and sensitivity_score columns reported per variant).",
        f"Substrate-specific subset: {len(ss):,} variants with |sensitivity - abundance| > 0.3 normalized units.",
        "",
        "Referenced in paper: Section 3.2 (functional heterogeneity), Section 3.6 (AlphaMissense stratified analysis), Figure 1 Panel C, Figure 2 Panels C–D, Table 4.",
        "",
        "Category thresholds (matching paper Section 3.6):",
        "  concordant-resistant   : abundance > 0.7 AND sensitivity > 0.7",
        "  concordant-sensitive   : abundance < 0.3 AND sensitivity < 0.5",
        "  thio-sensitive-stable  : abundance > 0.7 AND sensitivity < 0.5",
        "  paradoxical-resistant  : abundance < 0.3 AND sensitivity > 0.8",
        "  intermediate           : otherwise",
        "",
        "Sheets:",
        "  1) 'All paired variants' — complete joint set, 11 columns:",
        "       hgvs_pro           : HGVS protein-level variant identifier (e.g., p.Ala123Val).",
        "       WT, pos, mut       : Wild-type AA, 1-based position, mutant AA.",
        "       abundance_score    : Normalized VAMP-seq protein abundance (Suiter et al., 2020).",
        "       sensitivity_score  : Normalized thiopurine sensitivity (Suiter et al., 2020). Higher = more resistant.",
        "       delta              : sensitivity_score minus abundance_score.",
        "       abs_delta          : |delta|, used as the substrate-specificity magnitude.",
        "       category           : Biochemical category (see thresholds above).",
        "       am_score           : AlphaMissense pathogenicity score, 0 (benign) to 1 (pathogenic).",
        "       am_class           : AlphaMissense discrete classification (benign / ambiguous / pathogenic).",
        "",
        "  2) 'Substrate-specific subset' — same columns, restricted to rows with abs_delta > 0.3.",
        "  3) 'Category summary' — per-category N, mean abundance, mean sensitivity, mean AM score.",
        "  4) 'AM class per category' — cross-tabulation of category x AlphaMissense class counts.",
        "",
        "Data sources:",
        "  DMS: Suiter C.C., et al. (2020), MaveDB urn:mavedb:00000055-0-1 (combined), 00000055-a-1 (stability), 00000055-b-1 (sensitivity).",
        "  AlphaMissense: Cheng J., et al. (2023), precomputed scores for human canonical sequences (UniProt Q9NV35).",
    ]
    finalize_workbook(out_path, "Supplementary Table S3 — NUDT15 paired variants with substrate-specific annotation", readme_lines)
    print(f"[S3] Saved: {out_path}  (all={len(df):,}, substrate-specific={len(ss):,})")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("  Generating supplementary tables for PharmEpi benchmark paper")
    print("=" * 70)
    am_lookup = load_am()
    build_s1()
    build_s2(am_lookup)
    build_s3(am_lookup)
    print("\n[DONE] All supplementary tables generated at results/supplementary/")


if __name__ == "__main__":
    main()
